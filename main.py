import openpyxl
import matplotlib.pyplot as plt
import numpy as np
"""
This program calculates the time of the highest traffic in the telecommunications network based on the provided files.
    
File TIME.xlsx contains, expressed in seconds, the duration of calls recorded within one day.

File INT.xlsx It contains two columns of data. The first column lists individual minutes during the day, while the
    second column contains information about the number of calls that were recorded in a given minute of the day.    
"""


# adph
# fdmp
# fdmh
# tcbh


def calculate_mean(file_path):
    time_dataframe = openpyxl.load_workbook(file_path)
    time_data = time_dataframe['Sheet1']

    max_row = time_data.max_row
    max_column = time_data.max_column
    column_sums = [0] * max_column

    # Iterate through each row and add values to column sums
    for row in time_data.iter_rows(min_row=2, values_only=True):
        for col_index, value in enumerate(row):
            column_sums[col_index] += value

    # Calculate the mean for column
    column_means_m = [(sum_column / max_row) / 60 for sum_column in column_sums]
    time_dataframe.close()
    return column_means_m


def read_excel_file_int(file_path):
    int_dataframe = openpyxl.load_workbook(file_path)
    intensity_data = int_dataframe['Sheet1']
    points = []
    mean = calculate_mean('TIME.xlsx')

    # Assume that the first column contains first-category data and the second column contains second-category data
    for row in intensity_data.iter_rows(values_only=True):
        data_tuple = (row[0], row[1])   # min
        print("--" * 30)
        print(f"Values without means:   {data_tuple} (min)")
        data_tuple = (row[0], row[1] * mean[0])
        print(f"Values with means:      {data_tuple} (min)")

        points.append(data_tuple)
    int_dataframe.close()
    return points


def plot_data_adph():
    path_int = 'INT.xlsx'
    point = read_excel_file_int(path_int)

    x = list(map(lambda x1: x1[0], point))
    y = list(map(lambda x2: x2[1], point))

    plt.rc('grid', linestyle="-", color='black')
    plt.scatter(x, y)
    plt.xticks(np.arange(0, max(x) + 10, 60))
    plt.title("Peak Telecom Traffic Hour", fontdict={'family': 'serif', 'color': 'darkred', 'size': 16, })
    plt.xlabel("Time (min)")
    plt.ylabel("Call Counts (N)")
    plt.grid()
    plt.show()


plot_data_adph()
